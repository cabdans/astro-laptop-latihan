# -*- coding: utf-8 -*-
"""
Update LAPTOP_TANPA_LINK_AFFILIATE.xlsx SECARA AMAN.

ATURAN KERAS (jangan pernah dilanggar):
  1. DILARANG `del wb[sheet]` / `wb.remove(...)` / `Workbook()` baru lalu save ke path ini.
  2. DILARANG menulis apa pun ke kolom C (Link Affiliate) dan D (Note) pada baris
     yang sudah ada. Kolom itu MILIK CAHYA.
  3. Skrip ini hanya boleh: (a) menambah baris laptop baru di bawah, (b) melapor
     laptop yang sudah tidak ada di project.
  4. Selalu buat backup ber-timestamp sebelum menyimpan.

Cara pakai:
    python scripts/update_affiliate_xlsx.py            # laporan saja, tidak menyimpan
    python scripts/update_affiliate_xlsx.py --apply    # tambah baris baru + simpan
"""
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "LAPTOP_TANPA_LINK_AFFILIATE.xlsx"
BACKUP_DIR = ROOT / "backups" / "affiliate-xlsx"
REVIEW_DIR = ROOT / "src" / "pages" / "review"

KATEGORI_SHEET = {
    "budget": "Budget",
    "produktivitas": "Produktivitas",
    "gaming": "Gaming",
    "high-gaming": "High Gaming",
    "ultrabook": "Ultrabook",
}

KUNING = PatternFill("solid", fgColor="FFF7CC")
BODY = Font(name="Calibri", size=11)


def baca_project():
    """Kumpulkan (nama laptop, kategori) dari semua halaman review."""
    hasil = []
    for folder, sheet in KATEGORI_SHEET.items():
        d = REVIEW_DIR / folder
        if not d.is_dir():
            continue
        for f in sorted(d.glob("*.astro")):
            teks = f.read_text(encoding="utf-8", errors="ignore")
            m = re.search(r'title\s*=\s*["\']([^"\']+)["\']', teks)
            if not m:
                m = re.search(r'title\s*=\s*\{?["\']([^"\']+)["\']', teks)
            nama = m.group(1).strip() if m else f.stem
            hasil.append((nama, sheet, f.stem))
    return hasil


def kunci(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def main(apply_changes=False):
    if not XLSX.exists():
        sys.exit(f"File tidak ditemukan: {XLSX}")

    wb = load_workbook(XLSX)
    project = baca_project()

    # --- inventaris isi Excel saat ini (JANGAN diubah) -----------------
    ada = {}          # kunci nama -> (sheet, row)
    terisi = 0
    for sheet in KATEGORI_SHEET.values():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            nama = ws.cell(r, 1).value
            if not nama:
                continue
            ada[kunci(nama)] = (sheet, r)
            if ws.cell(r, 3).value:
                terisi += 1

    baru = [(n, s) for n, s, _ in project if kunci(n) not in ada]

    print(f"Baris di Excel   : {len(ada)}")
    print(f"Link sudah diisi : {terisi}   <-- WAJIB tetap utuh setelah skrip jalan")
    print(f"Laptop di project: {len(project)}")
    print(f"Perlu ditambah   : {len(baru)}")
    for n, s in baru:
        print(f"   + [{s}] {n}")

    hilang = [n for k, (s, r) in ada.items()
              for n in [wb[s].cell(r, 1).value]
              if k not in {kunci(x[0]) for x in project}]
    if hilang:
        print(f"\nAda di Excel tapi tidak di project ({len(hilang)}) "
              f"- TIDAK dihapus otomatis, cek manual:")
        for n in hilang:
            print(f"   ? {n}")

    if not apply_changes:
        print("\n(mode laporan. jalankan dengan --apply untuk menyimpan)")
        return

    if not baru:
        print("\nTidak ada yang perlu ditambah. File tidak disentuh.")
        return

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    cad = BACKUP_DIR / f"{datetime.now():%Y%m%d-%H%M%S}_LAPTOP_TANPA_LINK_AFFILIATE.xlsx"
    shutil.copy2(XLSX, cad)
    print(f"\nBackup dibuat: {cad.relative_to(ROOT)}")

    for nama, sheet in baru:
        ws = wb[sheet]
        r = ws.max_row + 1
        while ws.cell(r - 1, 1).value in (None, "") and r > 2:
            r -= 1
        ws.cell(r, 1, nama).font = BODY
        ws.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(r, 2, sheet).font = BODY
        c = ws.cell(r, 3)          # kolom link: DIBIARKAN KOSONG, tidak di-set None
        c.fill = KUNING
        c.font = BODY

    wb.save(XLSX)

    # --- verifikasi: link lama harus masih ada -------------------------
    cek = load_workbook(XLSX)
    sisa = sum(1 for s in KATEGORI_SHEET.values() if s in cek.sheetnames
               for r in range(2, cek[s].max_row + 1) if cek[s].cell(r, 3).value)
    print(f"Tersimpan. Link terisi sebelum={terisi} sesudah={sisa} "
          f"{'OK' if sisa >= terisi else '!!! DATA HILANG - RESTORE BACKUP !!!'}")


if __name__ == "__main__":
    main("--apply" in sys.argv)
